import io
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.views import View
from openpyxl import Workbook, load_workbook
from .models import Siswa, DataOrangTua


from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def download_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Siswa"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    headers = [
        'nisn', 'nik', 'nama_lengkap', 'jenis_kelamin', 'tempat_lahir',
        'tanggal_lahir', 'agama', 'golongan_darah', 'catatan_medis',
        'alamat_tinggal', 'status_siswa', 'tanggal_masuk', 'nama_ayah',
        'pekerjaan_ayah', 'no_hp_ayah', 'nama_ibu', 'pekerjaan_ibu',
        'no_hp_ibu', 'nama_wali', 'pekerjaan_wali', 'no_hp_wali'
    ]

    ws.append(headers)

    example_row = [
        '1234567890', '3201010101010001', 'Budi Utomo', 'L', 'Jakarta',
        '2000-01-15', 'Islam', 'O', '', 'Jl. Merdeka No. 1',
        'Aktif', '2020-07-01', 'Ahmad Utomo', 'Guru', '081234567890',
        'Siti Aminah', 'Ibu Rumah Tangga', '081234567891',
        'Wali 1', 'Wiraswasta', '081234567892'
    ]
    ws.append(example_row)

    # Styles
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    align_center = Alignment(horizontal='center', vertical='center')

    # Apply style to Header
    ws.row_dimensions[1].height = 25
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Format Example Row
    ws.row_dimensions[2].height = 20
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(name=font_family, size=10, color="6B7280", italic=True) # Gray out example row
        cell.border = thin_border
        if col_idx in [1, 2, 4, 6, 7, 8, 11, 12]:
            cell.alignment = align_center

    # Setup Data Validation (Dropdowns) for data rows (row 2 to 1000)
    # 1. Jenis Kelamin (Col D / 4)
    dv_jk = DataValidation(type="list", formula1='"L,P"', allow_blank=True)
    dv_jk.error ='Nilai harus L atau P'
    dv_jk.errorTitle = 'Input Tidak Valid'
    dv_jk.prompt = 'Pilih L (Laki-laki) atau P (Perempuan)'
    dv_jk.promptTitle = 'Jenis Kelamin'
    ws.add_data_validation(dv_jk)
    dv_jk.add("D2:D1000")

    # 2. Agama (Col G / 7)
    dv_agama = DataValidation(type="list", formula1='"Islam,Kristen,Katolik,Hindu,Buddha,Khonghucu"', allow_blank=True)
    dv_agama.error ='Pilih dari list agama resmi'
    dv_agama.errorTitle = 'Input Tidak Valid'
    dv_agama.prompt = 'Pilih Agama'
    dv_agama.promptTitle = 'Agama'
    ws.add_data_validation(dv_agama)
    dv_agama.add("G2:G1000")

    # 3. Golongan Darah (Col H / 8)
    dv_goldar = DataValidation(type="list", formula1='"A,B,AB,O"', allow_blank=True)
    dv_goldar.error ='Pilih golongan darah A, B, AB, atau O'
    dv_goldar.errorTitle = 'Input Tidak Valid'
    dv_goldar.prompt = 'Pilih Golongan Darah'
    dv_goldar.promptTitle = 'Golongan Darah'
    ws.add_data_validation(dv_goldar)
    dv_goldar.add("H2:H1000")

    # 4. Status Siswa (Col K / 11)
    dv_status = DataValidation(type="list", formula1='"Aktif,Lulus,Mutasi Keluar,Drop Out"', allow_blank=True)
    dv_status.error ='Pilih status siswa dari list'
    dv_status.errorTitle = 'Input Tidak Valid'
    dv_status.prompt = 'Pilih Status'
    dv_status.promptTitle = 'Status Siswa'
    ws.add_data_validation(dv_status)
    dv_status.add("K2:K1000")

    # Set column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="template_data_siswa.xlsx"'
    wb.save(response)
    return response

class SiswaImportView(LoginRequiredMixin, View):
    template_name = 'siswa/siswa_import.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        if 'import_file' not in request.FILES:
            messages.error(request, 'Tidak ada file yang diupload.')
            return redirect('siswa:import')

        file = request.FILES['import_file']
        if not file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Format file harus .xlsx atau .xls')
            return redirect('siswa:import')

        wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True)
        ws = wb.active

        created = 0
        errors = []
        header_map = {}

        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            header_map = {str(cell).lower().strip(): i for i, cell in enumerate(row) if cell}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[header_map.get('nisn', 0)]:
                    continue

                siswa, _ = Siswa.objects.update_or_create(
                    nisn=str(row[header_map.get('nisn', 0)]),
                    defaults={
                        'nik': row[header_map.get('nik', 1)] or '',
                        'nama_lengkap': row[header_map.get('nama_lengkap', 2)] or '',
                        'jenis_kelamin': row[header_map.get('jenis_kelamin', 3)] or 'L',
                        'tempat_lahir': row[header_map.get('tempat_lahir', 4)] or '',
                        'tanggal_lahir': row[header_map.get('tanggal_lahir', 5)],
                        'agama': row[header_map.get('agama', 6)] or '',
                        'golongan_darah': row[header_map.get('golongan_darah', 7)] or '',
                        'catatan_medis': row[header_map.get('catatan_medis', 8)] or '',
                        'alamat_tinggal': row[header_map.get('alamat_tinggal', 9)] or '',
                        'status_siswa': row[header_map.get('status_siswa', 10)] or 'Aktif',
                        'tanggal_masuk': row[header_map.get('tanggal_masuk', 11)],
                    }
                )

                DataOrangTua.objects.update_or_create(
                    siswa=siswa,
                    defaults={
                        'nama_ayah': row[header_map.get('nama_ayah', 12)] or '',
                        'pekerjaan_ayah': row[header_map.get('pekerjaan_ayah', 13)] or '',
                        'no_hp_ayah': row[header_map.get('no_hp_ayah', 14)] or '',
                        'nama_ibu': row[header_map.get('nama_ibu', 15)] or '',
                        'pekerjaan_ibu': row[header_map.get('pekerjaan_ibu', 16)] or '',
                        'no_hp_ibu': row[header_map.get('no_hp_ibu', 17)] or '',
                        'nama_wali': row[header_map.get('nama_wali', 18)] or '',
                        'pekerjaan_wali': row[header_map.get('pekerjaan_wali', 19)] or '',
                        'no_hp_wali': row[header_map.get('no_hp_wali', 20)] or '',
                    }
                )
                created += 1
            except Exception as e:
                errors.append(f'Baris {row_idx}: {str(e)}')

        if created > 0:
            messages.success(request, f'Berhasil mengimport {created} data siswa.')
        if errors:
            messages.warning(request, f'Terdapat {len(errors)} error: {", ".join(errors[:5])}')

        return redirect('siswa:index')