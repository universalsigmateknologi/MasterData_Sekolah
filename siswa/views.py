from django.shortcuts import render, redirect
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Q
from django.forms import inlineformset_factory
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Siswa, DataOrangTua
from .forms import SiswaForm, DataOrangTuaForm

DataOrangTuaFormSet = inlineformset_factory(
    Siswa, DataOrangTua, form=DataOrangTuaForm,
    extra=1, can_delete=False, can_delete_extra=False
)

from django.shortcuts import render
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Siswa
from akademik.models import Kelas, TahunAjaran
from kesiswaan.models import PenempatanKelas

def siswa_list(request):
    # Mengambil parameter GET untuk filter
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    jk_filter = request.GET.get('jenis_kelamin', '')
    kelas_filter = request.GET.get('kelas_filter', '')

    # Mendapatkan Tahun Ajaran yang sedang aktif sebagai acuan kelas
    active_ta = TahunAjaran.objects.filter(is_aktif=True).first()

    # Base queryset
    queryset = Siswa.objects.all().order_by('-created_at')

    # Logika Filter Berdasarkan Kelas (Hanya berlaku jika ada TA aktif)
    if active_ta:
        if kelas_filter == 'none':
            # Jika filter "Belum memiliki kelas", exclude siswa yang sudah ada di PenempatanKelas
            placed_ids = PenempatanKelas.objects.filter(
                tahun_ajaran=active_ta
            ).values_list('siswa_id', flat=True)
            queryset = queryset.exclude(id__in=placed_ids)
        elif kelas_filter:
            # Jika memilih kelas tertentu, HANYA tampilkan siswa yang statusnya AKTIF di kelas tersebut
            placed_ids = PenempatanKelas.objects.filter(
                kelas_id=kelas_filter,
                tahun_ajaran=active_ta,
                keterangan='Aktif' # <-- DITAMBAHKAN
            ).values_list('siswa_id', flat=True)
            queryset = queryset.filter(id__in=placed_ids)

    # Logika Filter Lainnya
    if search_query:
        queryset = queryset.filter(
            Q(nama_lengkap__icontains=search_query) |
            Q(nisn__icontains=search_query) |
            Q(nik__icontains=search_query)
        )
    if status_filter:
        queryset = queryset.filter(status_siswa=status_filter)
    if jk_filter:
        queryset = queryset.filter(jenis_kelamin=jk_filter)

    # Pagination
    paginator = Paginator(queryset, 10)
    page = request.GET.get('page', 1)

    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    # Logika pembuatan page_range (menghindari nomor halaman kebanyakan)
    current_page = page_obj.number
    page_range = []
    for num in paginator.page_range:
        if num <= 3 or num > paginator.num_pages - 3:
            page_range.append(num)
        elif abs(num - current_page) <= 2:
            page_range.append(num)

    context = {
        'siswa_list': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'page_range': page_range,
        # Nilai filter saat ini
        'search_query': search_query,
        'status_filter': status_filter,
        'jk_filter': jk_filter,
        'kelas_filter': kelas_filter,
        # Opsi dropdown
        'list_kelas': Kelas.objects.select_related('jurusan').order_by('tingkat', 'nama_kelas'),
    }
    
    return render(request, 'siswa/siswa_list.html', context)


class SiswaCreateView(LoginRequiredMixin, CreateView):
    model = Siswa
    template_name = 'siswa/siswa_form.html'
    form_class = SiswaForm
    success_url = reverse_lazy('siswa:index')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['ortu_form'] = DataOrangTuaFormSet(self.request.POST, self.request.FILES)
        else:
            context['ortu_form'] = DataOrangTuaFormSet()
        return context

    def form_valid(self, form):
        self.object = form.save()
        formset = DataOrangTuaFormSet(self.request.POST, self.request.FILES, instance=self.object)
        if formset.is_valid():
            formset.save()
            messages.success(self.request, 'Data siswa berhasil ditambahkan.')
            return redirect(self.success_url)
        return self.form_invalid(form)


class SiswaUpdateView(LoginRequiredMixin, UpdateView):
    model = Siswa
    template_name = 'siswa/siswa_form.html'
    pk_url_kwarg = 'pk'
    form_class = SiswaForm
    success_url = reverse_lazy('siswa:index')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['ortu_form'] = DataOrangTuaFormSet(self.request.POST, self.request.FILES, instance=self.object)
        else:
            context['ortu_form'] = DataOrangTuaFormSet(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        ortu_form = context['ortu_form']
        self.object = form.save()
        if ortu_form.is_valid():
            ortu_form.save()
            messages.success(self.request, 'Data siswa berhasil diperbarui.')
            return redirect('siswa:index')
        return self.form_invalid(form)


class SiswaDeleteView(LoginRequiredMixin, DeleteView):
    model = Siswa
    template_name = 'siswa/siswa_confirm_delete.html'
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('siswa:index')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Data siswa berhasil dihapus.')
        return response


@login_required
def export_siswa_excel(request):
    queryset = Siswa.objects.all().select_related('orang_tua')
    search = request.GET.get('search')
    status = request.GET.get('status')
    jenis_kelamin = request.GET.get('jenis_kelamin')
    kelas_filter = request.GET.get('kelas_filter')

    active_ta = TahunAjaran.objects.filter(is_aktif=True).first()

    if active_ta and kelas_filter:
        if kelas_filter == 'none':
            placed_ids = PenempatanKelas.objects.filter(
                tahun_ajaran=active_ta
            ).values_list('siswa_id', flat=True)
            queryset = queryset.exclude(id__in=placed_ids)
        else:
            placed_ids = PenempatanKelas.objects.filter(
                kelas_id=kelas_filter,
                tahun_ajaran=active_ta
            ).values_list('siswa_id', flat=True)
            queryset = queryset.filter(id__in=placed_ids)

    if search:
        queryset = queryset.filter(
            Q(nama_lengkap__icontains=search) |
            Q(nisn__icontains=search) |
            Q(nik__icontains=search)
        )
    if status:
        queryset = queryset.filter(status_siswa=status)
    if jenis_kelamin:
        queryset = queryset.filter(jenis_kelamin=jenis_kelamin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Siswa"
    
    ws.views.sheetView[0].showGridLines = True

    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1F2937")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="4B5563")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="374151")
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws['A1'] = "LAPORAN DATA BUKU INDUK SISWA"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 25
    
    current_time = timezone.localtime(timezone.now()).strftime('%d-%m-%Y %H:%M:%S')
    ws['A2'] = f"Dicetak pada: {current_time} | Total Data: {queryset.count()}"
    ws['A2'].font = subtitle_font
    ws.row_dimensions[2].height = 18

    filters_desc = []
    if search:
        filters_desc.append(f"Pencarian: '{search}'")
    if status:
        filters_desc.append(f"Status: {status}")
    if jenis_kelamin:
        jk_str = "Laki-laki" if jenis_kelamin == "L" else "Perempuan"
        filters_desc.append(f"Jenis Kelamin: {jk_str}")
    if kelas_filter:
        if kelas_filter == 'none':
            filters_desc.append("Kelas: Belum Memiliki Kelas")
        else:
            kelas_nama = Kelas.objects.filter(id=kelas_filter).first()
            if kelas_nama:
                filters_desc.append(f"Kelas: {kelas_nama.nama_kelas}")
    
    if filters_desc:
        ws['A3'] = f"Filter aktif: {', '.join(filters_desc)}"
    else:
        ws['A3'] = "Filter aktif: Semua Data"
    ws['A3'].font = Font(name=font_family, size=9, italic=True, color="6B7280")
    ws.row_dimensions[3].height = 15

    headers = [
        "No", "NISN", "NIK", "Nama Lengkap", "L/P", "Tempat Lahir", 
        "Tanggal Lahir", "Agama", "Golongan Darah", "Alamat Tinggal", 
        "Status Siswa", "Tanggal Masuk", "Nama Ayah", "No. HP Ayah", 
        "Nama Ibu", "No. HP Ibu", "Nama Wali", "No. HP Wali"
    ]
    
    header_row = 5
    ws.row_dimensions[header_row].height = 28
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    current_row = header_row + 1
    for idx, siswa in enumerate(queryset, 1):
        jk = "L" if siswa.jenis_kelamin == "L" else "P"
        
        nama_ayah = ""
        no_hp_ayah = ""
        nama_ibu = ""
        no_hp_ibu = ""
        nama_wali = ""
        no_hp_wali = ""
        
        try:
            if hasattr(siswa, 'orang_tua'):
                ortu = siswa.orang_tua
                nama_ayah = ortu.nama_ayah
                no_hp_ayah = ortu.no_hp_ayah
                nama_ibu = ortu.nama_ibu
                no_hp_ibu = ortu.no_hp_ibu
                nama_wali = ortu.nama_wali
                no_hp_wali = ortu.no_hp_wali
        except Exception:
            pass

        row_data = [
            idx,
            siswa.nisn or "-",
            siswa.nik or "-",
            siswa.nama_lengkap,
            jk,
            siswa.tempat_lahir,
            siswa.tanggal_lahir.strftime('%d-%m-%Y') if siswa.tanggal_lahir else "-",
            siswa.agama or "-",
            siswa.golongan_darah or "-",
            siswa.alamat_tinggal,
            siswa.status_siswa,
            siswa.tanggal_masuk.strftime('%d-%m-%Y') if siswa.tanggal_masuk else "-",
            nama_ayah or "-",
            no_hp_ayah or "-",
            nama_ibu or "-",
            no_hp_ibu or "-",
            nama_wali or "-",
            no_hp_wali or "-"
        ]
        
        ws.row_dimensions[current_row].height = 20
        is_even = (idx % 2 == 0)
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            
            if col_idx in [1, 2, 3, 5, 7, 9, 11, 12, 14, 16, 18]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            if is_even:
                cell.fill = zebra_fill
                
        current_row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col[4:]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['J'].width = 40

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Data_Siswa_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response